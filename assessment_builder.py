"""
assessment_builder.py
Word Assessment for WFA — PDF pupil sheets + Excel marking grid.
"""

import io, json, os, requests, base64, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import simpleSplit

from word_bank import WORD_BANK

# ── Palette ───────────────────────────────────────────────────────────────────
BLUE   = (23/255, 152/255, 211/255)   # WFA #1798d3
WHITE  = (1, 1, 1)
NAVY   = (0.10, 0.22, 0.43)
LGREY  = (0.95, 0.95, 0.95)
DGREY  = (0.27, 0.27, 0.27)
BLACK  = (0, 0, 0)

XL_BLUE  = "FF1798D3"
XL_LBLUE = "FFDEECF8"
XL_WHITE = "FFFFFFFF"
XL_GREY  = "FFF2F2F2"
XL_BLACK = "FF000000"

# ── Word sets ─────────────────────────────────────────────────────────────────
FOCUS_Y2 = [
    "friend", "people", "many", "it's", "break", "climb",
    "prove", "sugar", "whole", "because", "beautiful"
]

def _words_for_year(yr):
    return [w for w, y, *_ in WORD_BANK if y == yr]

def _all_sections():
    return {
        "Y1/Y2": ("Commonly misspelled Y1 / Y2 words", FOCUS_Y2),
        "Y3":    ("Year 3 key spellings",               _words_for_year("3")),
        "Y4":    ("Year 4 key spellings",               _words_for_year("4")),
    }


# ── PDF pupil sheet ───────────────────────────────────────────────────────────

def _draw_page_header(c, W, H, pupil_name, week_ref, page_label):
    """Blue header bar with name and week ref. Returns y below header."""
    HDR = 14 * mm
    c.setFillColorRGB(*BLUE)
    c.rect(0, H - HDR, W, HDR, fill=1, stroke=0)
    c.setFillColorRGB(*WHITE)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(10 * mm, H - HDR + (HDR - 11) / 2, pupil_name)
    c.setFont("Helvetica", 8)
    c.drawRightString(W - 10 * mm, H - HDR + (HDR - 8) / 2,
                      f"Word Assessment  ·  {week_ref}  ·  {page_label}")
    return H - HDR


def build_word_assessment_pdf(pupils, sections, cloze_lookup, week_ref=""):
    """
    Returns PDF bytes. One or more pages per pupil (name on every page).
    sections: list of (label, [words])
    cloze_lookup: {word_lower: sentence_with_blanks}
    """
    buf = io.BytesIO()
    W, H = A4
    M    = 12 * mm
    UW   = W - 2 * M

    # Row layout
    ROW_H    = 9   * mm   # 0.9cm per row
    SEC_H    = 6   * mm   # section header
    FIRST_Y  = H - 14 * mm - 5 * mm
    SUBSEQ_Y = H - 14 * mm - 5 * mm

    c = canvas.Canvas(buf, pagesize=A4)

    # Column widths
    NUM_W  = 8  * mm
    MARK_W = 8  * mm   # marking box column
    SENT_W = UW - NUM_W - MARK_W

    all_words_flat = []
    for label, words in sections:
        all_words_flat.extend(words)

    for pupil in pupils:
        name = pupil.get("first", "") + " " + (pupil.get("last") or "")
        name = name.strip()

        # Build flat list of items: ("section", label) or ("word", n, word, sentence)
        items = []
        counter = 1
        for label, words in sections:
            items.append(("section", label))
            for word in words:
                sent = cloze_lookup.get(word.lower(), f"Write the word: {'_' * 30}.")
                items.append(("word", counter, word, sent))
                counter += 1

        # Paginate
        page_num = 0
        item_idx = 0
        total_pages = None   # filled after first pass

        # First pass to count pages (then draw)
        def paginate(items):
            pages = []
            current = []
            cy = FIRST_Y
            for item in items:
                need = SEC_H if item[0] == "section" else ROW_H
                if current and cy - need < M:
                    pages.append(current)
                    current = []
                    cy = SUBSEQ_Y
                current.append(item)
                cy -= need
            if current:
                pages.append(current)
            return pages

        pages = paginate(items)
        n_pages = len(pages)

        for pg_idx, page_items in enumerate(pages):
            pg_label = f"Page {pg_idx + 1} of {n_pages}"
            top_y = _draw_page_header(c, W, H, name, week_ref, pg_label)
            cy = top_y - 6 * mm

            for item in page_items:
                if item[0] == "section":
                    _, label = item
                    # Section header bar
                    c.setFillColorRGB(*LGREY)
                    c.rect(M, cy - SEC_H, UW, SEC_H, fill=1, stroke=0)
                    c.setFillColorRGB(*NAVY)
                    c.setFont("Helvetica-Bold", 8.5)
                    baseline = cy - SEC_H + (SEC_H - 8.5) / 2
                    c.drawString(M + 3 * mm, baseline, label)
                    cy -= SEC_H

                else:
                    _, num, word, sentence = item
                    row_top = cy
                    row_bot = cy - ROW_H

                    # Alternating row tint
                    if num % 2 == 0:
                        c.setFillColorRGB(0.97, 0.97, 0.97)
                        c.rect(M, row_bot, UW, ROW_H, fill=1, stroke=0)

                    # Light top divider
                    c.setStrokeColorRGB(0.82, 0.82, 0.82)
                    c.setLineWidth(0.3)
                    c.line(M, row_top, M + UW, row_top)

                    # Number — vertically centred
                    mid_y = row_bot + ROW_H / 2
                    c.setFillColorRGB(*NAVY)
                    c.setFont("Helvetica-Bold", 8)
                    c.drawCentredString(M + NUM_W / 2, mid_y - 8 * 0.35, str(num))

                    # Writing line — stops before marking box
                    line_y = row_bot + 2.2 * mm
                    c.setStrokeColorRGB(0.50, 0.50, 0.50)
                    c.setLineWidth(0.5)
                    c.line(M + NUM_W, line_y, M + UW - MARK_W - 2 * mm, line_y)

                    # Sentence — extend any underscore run to 35 for writing space
                    import re as _re
                    display_sent = _re.sub(r'_+', '_' * 35, sentence)
                    c.setFillColorRGB(*BLACK)
                    c.setFont("Helvetica", 8)
                    sent_x  = M + NUM_W + 2 * mm
                    sent_bl = line_y + 1.2 * mm
                    c.drawString(sent_x, sent_bl, display_sent)

                    # Marking box — right-aligned, vertically centred
                    BOX_SZ = 6 * mm
                    box_x  = M + UW - MARK_W + (MARK_W - BOX_SZ) / 2
                    box_y  = row_bot + (ROW_H - BOX_SZ) / 2
                    c.setStrokeColorRGB(0.35, 0.35, 0.35)
                    c.setLineWidth(0.7)
                    c.rect(box_x, box_y, BOX_SZ, BOX_SZ, fill=0, stroke=1)

                    cy -= ROW_H

            # Bottom border
            c.setStrokeColorRGB(0.80, 0.80, 0.80)
            c.setLineWidth(0.4)
            c.line(M, cy, M + UW, cy)

            c.showPage()

    c.save()
    buf.seek(0)
    return buf.read()


# ── Excel marking grid ────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, value, bg_hex, fg_hex="FFFFFFFF", rotate=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg_hex)
    c.font = Font(name="Calibri", bold=True, color=fg_hex, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True,
                            text_rotation=45 if rotate else 0)
    c.border = _thin()
    return c

def _data_cell(ws, row, col, value="", bold=False, bg_hex=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _thin()
    if bg_hex:
        c.fill = PatternFill("solid", fgColor=bg_hex)
    return c


def build_word_assessment_excel(pupils, sections):
    """
    Two-tab Excel: Marking Grid + Results (blank, filled after import).
    sections: list of (label, [words])
    Returns bytes.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_words = []
    section_spans = []   # (label, start_idx, end_idx)
    for label, words in sections:
        start = len(all_words)
        all_words.extend(words)
        section_spans.append((label, start, len(all_words) - 1))

    FIXED = 3   # No. | First | Last

    # ── Tab 1: Marking Grid ───────────────────────────────────────────────
    mg = wb.create_sheet("Marking Grid")

    # Row 1: section labels (merged across word columns)
    for lbl in ["No.", "First", "Last"]:
        col = ["No.", "First", "Last"].index(lbl) + 1
        _hdr_cell(mg, 1, col, lbl, XL_BLUE[2:])

    for label, start, end in section_spans:
        sc = FIXED + start + 1
        ec = FIXED + end + 1
        c = _hdr_cell(mg, 1, sc, label, XL_LBLUE[2:], fg_hex="FF1A3C6E")
        if sc < ec:
            mg.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

    # Row 2: individual word headers (rotated)
    for i, word in enumerate(all_words):
        col = FIXED + i + 1
        _hdr_cell(mg, 2, col, word, XL_BLUE[2:], rotate=True)
        mg.column_dimensions[get_column_letter(col)].width = 5

    mg.column_dimensions["A"].width = 4
    mg.column_dimensions["B"].width = 12
    mg.column_dimensions["C"].width = 14
    mg.row_dimensions[1].height = 20
    mg.row_dimensions[2].height = 80

    # Pupil rows
    for r_idx, p in enumerate(pupils):
        row = r_idx + 3
        alt = XL_GREY[2:] if r_idx % 2 == 0 else None
        _data_cell(mg, row, 1, r_idx + 1, align="center", bg_hex=alt)
        _data_cell(mg, row, 2, p.get("first", ""), bg_hex=alt)
        _data_cell(mg, row, 3, p.get("last", ""),  bg_hex=alt)
        for ci in range(len(all_words)):
            c = mg.cell(row=row, column=FIXED + ci + 1, value="")
            c.border = _thin()
            if alt:
                c.fill = PatternFill("solid", fgColor=alt)

    mg.freeze_panes = "D3"

    # Print setup
    mg.page_setup.orientation = "landscape"
    mg.page_setup.paperSize   = mg.PAPERSIZE_A4
    mg.page_setup.fitToWidth  = 1
    mg.page_setup.fitToHeight = 0
    mg.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

    # ── Tab 2: Results (populated after import) ───────────────────────────
    rs = wb.create_sheet("Results")
    for col, hdr in enumerate(["First", "Last", "Words correct", "Words incorrect",
                                "% correct", "Notes"], 1):
        _hdr_cell(rs, 1, col, hdr, XL_BLUE[2:])
    for r_idx, p in enumerate(pupils):
        row = r_idx + 2
        alt = XL_GREY[2:] if r_idx % 2 == 0 else None
        _data_cell(rs, row, 1, p.get("first", ""), bg_hex=alt)
        _data_cell(rs, row, 2, p.get("last", ""),  bg_hex=alt)
        for col in range(3, 7):
            c = rs.cell(row=row, column=col, value="")
            c.border = _thin()
            if alt:
                c.fill = PatternFill("solid", fgColor=alt)
    rs.column_dimensions["A"].width = 12
    rs.column_dimensions["B"].width = 14
    for col in ["C","D","E","F"]:
        rs.column_dimensions[col].width = 14
    rs.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Cloze generation (gap-fill via Claude API) ────────────────────────────────

def generate_missing_cloze(words, api_key):
    """
    Call Claude API to generate cloze sentences for words missing from the bank.
    Returns {word_lower: sentence} dict.
    """
    word_list = "\n".join(f"- {w}" for w in words)
    prompt = (
        "You are helping a primary school teacher create a spelling test for Year 3 and 4 "
        "learners (aged 7-9) in England.\n\n"
        "For each word below, write ONE short cloze sentence. Rules:\n"
        "- Replace the target word with _______________ (15 underscores)\n"
        "- Sentences must be simple enough for a 7-9 year old to understand\n"
        "- Sentences must make the word's meaning clear from context\n"
        "- Keep sentences to one clause where possible\n"
        "- British English spelling throughout\n"
        "- Return ONLY valid JSON: object where each key is the word and each value is the sentence\n"
        "- No preamble, no markdown fences, no extra text\n\n"
        f"Words:\n{word_list}"
    )
    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"},
        json={"model": "claude-sonnet-4-20250514", "max_tokens": 2000,
              "messages": [{"role": "user", "content": prompt}]},
        timeout=30,
    )
    if resp.status_code == 200:
        text = resp.json()["content"][0]["text"].strip()
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        return {k.lower(): v for k, v in json.loads(text).items()}
    return {}
